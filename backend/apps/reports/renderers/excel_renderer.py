"""
Excel Renderer — Generates .xlsx files using openpyxl.
"""

import io

from .base import BaseRenderer


class ExcelRenderer(BaseRenderer):
    """Render report as Excel (.xlsx) file."""

    @property
    def content_type(self):
        return 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    @property
    def file_extension(self):
        return 'xlsx'

    def render(self) -> bytes:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl")

        wb = Workbook()
        ws = wb.active
        import re
        ws.title = re.sub(r'[\\/*?:\[\]]', '', self.get_title())[:31].strip() or 'Report'

        columns = self.get_columns()
        rows = self.get_rows()
        aggregations = self.get_aggregations()

        # Theme colors
        from .html_renderer import THEME_CSS
        theme = THEME_CSS.get(self.get_theme(), THEME_CSS['corporate_blue'])

        header_fill = PatternFill(
            start_color=theme['header_bg'].lstrip('#'),
            end_color=theme['header_bg'].lstrip('#'),
            fill_type='solid',
        )
        header_font = Font(
            name='Segoe UI', size=10, bold=True,
            color=theme['header_text'].lstrip('#'),
        )
        alt_fill = PatternFill(
            start_color=theme['row_alt'].lstrip('#'),
            end_color=theme['row_alt'].lstrip('#'),
            fill_type='solid',
        )
        total_fill = PatternFill(
            start_color=theme['total_bg'].lstrip('#'),
            end_color=theme['total_bg'].lstrip('#'),
            fill_type='solid',
        )
        total_font = Font(
            name='Segoe UI', size=10, bold=True,
            color=theme['total_text'].lstrip('#'),
        )
        body_font = Font(name='Segoe UI', size=9)
        thin_border = Border(
            left=Side(style='thin', color=theme['border'].lstrip('#')),
            right=Side(style='thin', color=theme['border'].lstrip('#')),
            top=Side(style='thin', color=theme['border'].lstrip('#')),
            bottom=Side(style='thin', color=theme['border'].lstrip('#')),
        )

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(columns), 1))
        title_cell = ws.cell(row=1, column=1, value=self.get_title())
        title_cell.font = Font(name='Segoe UI', size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='left')

        if self.get_subtitle():
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(columns), 1))
            sub_cell = ws.cell(row=2, column=1, value=self.get_subtitle())
            sub_cell.font = Font(name='Segoe UI', size=10, italic=True, color='666666')

        start_row = 4

        # Header row
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=col.get('label', col['path']))
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal=col.get('alignment', 'left'),
                vertical='center',
                wrap_text=True,
            )
            cell.border = thin_border

            # Auto column width
            width = col.get('width')
            if width:
                ws.column_dimensions[get_column_letter(col_idx)].width = max(width / 7, 12)
            else:
                ws.column_dimensions[get_column_letter(col_idx)].width = max(len(str(col.get('label', ''))), 12)

        # Data rows
        for row_idx, row_data in enumerate(rows, start_row + 1):
            for col_idx, col in enumerate(columns, 1):
                value = row_data.get(col['path'])
                formatted = self.format_value(value, col) if value is not None else ''

                # Try to preserve numeric values
                cell_value = value if isinstance(value, (int, float)) else formatted
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                cell.font = body_font
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal=col.get('alignment', 'left'),
                    vertical='top',
                )

                # Alternate row fill
                if (row_idx - start_row) % 2 == 0:
                    cell.fill = alt_fill

        # Aggregation footer
        if aggregations:
            total_row = start_row + 1 + len(rows)
            for col_idx, col in enumerate(columns, 1):
                if col_idx == 1:
                    cell = ws.cell(row=total_row, column=1, value='TOTALS')
                else:
                    alias = col['path'].lower().replace(' ', '_').replace('-', '_')
                    val = aggregations.get(alias)
                    cell = ws.cell(row=total_row, column=col_idx, value=val if val else '')

                cell.fill = total_fill
                cell.font = total_font
                cell.border = thin_border

        # Freeze header row
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
